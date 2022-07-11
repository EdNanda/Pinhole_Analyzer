"""
Created on Thu May 14 14:00:16 2020

@author: Edgar Nandayapa
"""

import numpy as np
import cv2
from glob import glob
import matplotlib.pyplot as plt
from astropy.modeling import models, fitting
from PIL import Image, ImageDraw, ImageFont, ImageOps
import os
import warnings
import sys
import openpyxl



## This collects and lists all microscopic images in a folder
def collect_images(sample_path):
    files = []
    img_array = []
    
    ## Order files by creation time
    # for img in sorted(glob(sample_path+"\\*"),key=os.path.getmtime):
    #     ## Check if they are mic pictures (filename should contain "_X" & "_Y")
    #     if "_X" and "_Y" in img:
    #         if not "pinhole" in img:
    #             files.append(img)
    #             ocv_img = cv2.imread(img,1) ##OpenCV read image
    #             img_array.append(ocv_img)
    #         else:
    #             continue

    for img in sorted(glob(sample_path+"\\*"),key=os.path.getmtime)[:25]:
        files.append(img)
        ocv_img = cv2.imread(img,1) ##OpenCV read image
        img_array.append(ocv_img)
            
    return files, img_array

## Identify and quantify black spots on images. Magic happens at cv2.threshold
# def count_defect_pinholes(imgs_array,coff):
def count_defect_pinholes(imgs_array):
    number_contours = []
    average_area = []
    pinhole_positions = []
    fit_array = 0

    for c,ima in enumerate(imgs_array):
        # print(c)
        zoom_factor = 7.9# px/um if taken with 100x objective
        ## Open image and set to grayscal
        gray = cv2.cvtColor(ima, cv2.COLOR_RGB2GRAY)
        
        fitted = quantify_image(gray) 
        
        if c == 0:
            fit_array = fitted
        else:
            # print(fit_array.shape,fitted.shape)
            fit_array = np.dstack((fit_array,fitted))
        
        bright = np.max(fitted)+1
        
        gray = gray-fitted+bright
        gray = gray.astype(np.uint8)
        
        coff = bright-25
        # print(f"  {bright}")
        ## For uneven lighting use cv2.adaptiveThreshold
        _,thresh = cv2.threshold(gray,coff,255,cv2.THRESH_BINARY_INV) #dark defects
        # _,thresh = cv2.threshold(gray,210,255,cv2.THRESH_BINARY) #light defects
        ### To fine tune, change the last number
        # thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 51, 20)
        
        ## Debugging
        # plt.imshow(thresh,'gray') #Draws image of pinholes
        # plt.show()

        ## Find contours/defects on images
        pinholes_list,_ = cv2.findContours(thresh, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
        # pinholes_list,_ = cv2.findContours(thresh, cv2.RETR_FLOODFILL, cv2.CHAIN_APPROX_SIMPLE)    
        ## Measure the area of the defects
        area = []
        contours = []
        for pinh in pinholes_list:
            ph_area = cv2.contourArea(pinh)
            if ph_area > 1:# and cv2.contourArea(cont) < 9000:
                area.append(ph_area)
                contours.append(pinh)
            else:
                continue

        ## Color defects depending if light or dark

        color = (0,255,0)#in yellow

        ## Necessary to save an image of pinholes later on
        pinhole_image = cv2.drawContours(ima, contours,-1,color,5)
        pinhole_positions.append(pinhole_image)
        # plt.imshow(pinhole_image,'gray') #Draws image of pinholes
        # plt.show()

        ## Count pinholes, calculate average size and stdev
        if area[:]:
            pinhole_count = len(contours)
            pinhole_sizes = np.mean(area)/zoom_factor
            # pinhole_error = np.std(area)
        else:
            pinhole_count = 0
            pinhole_sizes = 0
            # pinhole_error = 0

        number_contours.append(pinhole_count)

        if  pinhole_sizes > 5000:
            average_area.append(0)
        else:
            average_area.append(round(pinhole_sizes,1))

    return fit_array, number_contours, average_area, pinhole_positions

def quantify_image(file): ## For polynomial fitting
    ## Get image size
    xs,ys = file.shape
    
    ## This gives the position to each element in the matrix
    y, x = np.mgrid[:ys, :xs]
    z = file
    
    fit = resize_matrix(x,y,xs, ys, z)
    return fit

##Resize matrix to smaller size (to speed up fitting calculation)
def resize_matrix(x,y,xs,ys,z,dim=100): ## For polynomial fitting
    # dim = 100
    
    ## Make a reduced, equally-spaced sample of the fitting 
    sim_x = np.linspace(0,xs-1,dim).astype("int")
    sim_y = np.linspace(0,ys-1,dim).astype("int")
    
    mat_z = np.zeros([dim,dim]).astype("uint8")
    
    for ci,i in enumerate(sim_x):
        for cj,j in enumerate(sim_y):
            mat_z[ci,cj] = z[i,j]
    
    mat_x,mat_y =np.meshgrid(sim_x,sim_y)
    
    fit = fit_polynomial_to_matrix(mat_x, mat_y, mat_z, x, y)
    
    return fit


def fit_polynomial_to_matrix(mat_x,mat_y,mat_z,x,y): ## For polynomial fitting
    ##Fit reduced matrix
    deg = 4
    p_init = models.Polynomial2D(degree=deg)
    fit_p = fitting.LevMarLSQFitter()
    
    # print(f"1- {time()-start}")
    with warnings.catch_warnings():
        # Ignore model linearity warning from the fitter
        warnings.simplefilter('ignore')
        p = fit_p(p_init, mat_x, mat_y, mat_z)

    
    fit = p(x,y).astype("uint8")
    
    return fit

def save_image_with_pinholes(imgs_array, images, pics_side):
    total_pics = pics_side**2
    for n in list(range(total_pics)):
        cv2.imwrite(imgs_array[n].split(".")[0]+"_pinhole.png", images[n])



### Plotting functions

## This makes bubble plots
def plotting_bubble(grid_steps, defs, sample_name, def_path):#defs,fPath,sName):
    ## Identify if data of pinhole area or count
    if isinstance(defs[0],float):
        typeG = "area"
    else:
        typeG = "counts"

    ## Make a matrix  with the number of places where pictures were taken
    def_matrix = np.zeros([grid_steps,grid_steps])
    plt.cla()

    ## Make coordinates for the bubble positions
    cell_posit = []
    for i in range(grid_steps)[::-1]:
        for j in range(grid_steps):
            cell_posit.append((j,i))

    ## Place bubbles on the right coordinate
    for i,d in list(enumerate(cell_posit)):
        def_matrix[d] = defs[i]
        if defs[i] == 0:
            plt.scatter(cell_posit[i][0]+0.5,cell_posit[i][1]+0.5, s=0.001, color="c")
        else:
            plt.scatter(cell_posit[i][0]+0.5,cell_posit[i][1]+0.5,s=defs[i]/max(defs)*1700, color="c")  ## 1700 to normalize the maximum bubble radius to one square
        plt.text(cell_posit[i][0]+0.5,cell_posit[i][1]+0.5,defs[i],fontsize=8, horizontalalignment='center',
                  verticalalignment='center')

    ## Configuration of plot
    plt.xlim(0, grid_steps)
    plt.ylim(0, grid_steps)
    plt.grid()
    plt.xticks(list(range(grid_steps)),[])
    plt.yticks(list(range(grid_steps)),[])
    plt.gca().set_aspect('equal', adjustable='box')

    plt.xlabel(sample_name+" "+typeG)
    plt.savefig(f"{def_path}\\00_{sample_name}_pinhole_bubble_{typeG}",dpi=300,bbox_inches='tight',pad_inches=0)
    plt.close()
    
    
    
def boxplot_all_samples(names, data, path):
    if isinstance(data[1][0],float):
        typeG = "area"
    else:
        typeG = "counts"

    figaP, axaP = plt.subplots()
    axaP.set_title(f"All samples: {typeG}")

    if True:## Plot in Log scale
        #Setting y-axis to log leads to outliers
        log_data = []
        for d in data:
            log_data.append(np.ma.log10(d).filled(0)) #np. masked&filled with 0
    
        axaP.boxplot(log_data, showmeans=True)
    
        axaP.set_yticks(np.arange(0, 5))
        axaP.set_yticklabels(10.0**np.arange(0, 5))
        typeG = "log_"+typeG
        
    else: ##Plot with normal scale
        axaP.boxplot(data,showmeans=True)
        
    axaP.set_xticklabels(names)
    axaP.grid(axis="y", linestyle="--")
    plt.xticks(rotation=90)

    plt.savefig(f"{path}\\pinhole_boxplot_{typeG}",dpi=300)

    plt.close()



def results_to_csv_file(name_arr, count_arr, area_arr, path):
    n = len(name_arr)
    
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    
    ## Place data in an excel sheet
    ws.cell(row=1, column=1).value = "Number of pinholes"
    ws.cell(row=n+3, column=1).value = "Area of pinholes"
    
    for r, name in enumerate(name_arr):
        ws.cell(row=r+2, column=1).value = name
        ws.cell(row=n+r+4, column=1).value = name
        
        for c,_ in enumerate(count_arr[r]):
            ws.cell(row=r+2, column=c+2).value = count_arr[r][c]
            ws.cell(row=n+r+4, column=c+2).value = area_arr[r][c]
        
    wb.save(filename=path+'pinhole_results.xlsx')



## Build collage of all microscopic images
def build_collage(path, files):
    
    # Open all images
    img_obj = []
    for f in files:
        img = Image.open(f)
        grey = ImageOps.grayscale(img)
        img_obj.append(grey)

    # Coordinates of areas of interest
    xsize, ysize = img_obj[0].size
    i_size = xsize*1//3
    sp = 1 #spacing between images
    coords = (i_size, ysize//2-i_size//2, 2*i_size-sp , (ysize//2)+(i_size//2)-sp)

    # Crop all images to the coordinates chosen
    cropp_img = []
    for c in img_obj:
        cropp_img.append(c.crop(coords))

    # Calculate the number of sides of the matrix(image)
    grid = int(np.sqrt(len(files)))
    composed_image = Image.new('RGB', (i_size*grid, i_size*grid))   #Creates a new/blank image

    positions = []
    for m in list(range(grid)):
        for n in list(range(grid)):
            positions.append((n*i_size, m*i_size))
    # print(positions)

    #Stitch cropped images together
    for c, ci in enumerate(cropp_img):
        composed_image.paste(ci, positions[c])

    sample_name = path.rsplit('\\',2)[-2]
    end_path = path

    ##Add text with sample name
    font = ImageFont.truetype("arial.ttf", 20)
    ImageDraw.Draw(composed_image).text((0,0),sample_name,(0,0,0),font=font)

    #Save image to same folder where original is located
    # print(end_path+sample_name+"_defects_coax_composed.png")
    composed_image.save(end_path+"00_"+sample_name+"_pinhole_collage.png")





## SETUP: Location of sample folders and number of pictures taken per side
# folder_path = "D:\\Seafile\\LEXT_SolarCell\\"
# folder_path = "C:\\Users\\HYDN02\\Downloads\\Solar Cell\\"
# folder_path = "D:\\Seafile\\IJPrinting_Edgar-Florian\\20210726_3cat-tosy\\20210730_Lext_Tosy\\"
folder_path = "C:\\Users\\HYDN02\\Seafile\\IJPrinting_Edgar-Florian\\20220615_NMP\\"
pics_per_side = 5
# threshold_val = 155 # Choose from 0 to 255 (~150 best, higher = more sensitive)

## List all samples (folders) inside folder_path
samples_path = glob(folder_path+"\\*\\")

## initialize arrays
name_arr    = []
count_array = []
area_array  = []
fit_collect = 0

if not samples_path:
    print("Error: Select a valid folder_path")
    sys.exit()

for sample in samples_path[:]:
    ## Get the sample name
    sample_name = sample.split("\\")[-2]
    name_arr.append(sample_name)
    print(sample_name)
    
    ## Collect all microscopic images per sample & read with OpenCV
    images, img_array = collect_images(sample)

    ## Identify pinholes in image
    # fitted, pinhole_count, pinhole_area, images_with_contours = count_defect_pinholes(img_array,threshold_val)
    fitted, pinhole_count, pinhole_area, images_with_contours = count_defect_pinholes(img_array)
    # fit_collect.append(fitted)
    try:
        fit_collect = np.dstack((fit_collect, fitted))
    except:
        fit_collect = fitted
        
    # print(fit_collect.shape)

    count_array.append(pinhole_count)
    area_array.append(pinhole_area)
    
    ## Build collage of all microscopic images
    build_collage(sample, images)

    ## Save images with demarked pinholes
    save_image_with_pinholes(images, images_with_contours, pics_per_side)
        
    ## Plot distribution pinholes across the substrate
    plotting_bubble(pics_per_side, pinhole_area, sample_name, sample)
    plotting_bubble(pics_per_side, pinhole_count, sample_name, sample)
    


## Save excel file containing all results
results_to_csv_file(name_arr, count_array, area_array, folder_path)

## Boxplot comparing pinhole count & area of all samples
boxplot_all_samples(name_arr,count_array,folder_path)
boxplot_all_samples(name_arr,area_array,folder_path)

print("done!")


## For testing:  Plotting of fitted surface 
fc_std = np.std(fit_collect,axis=2)
fc_mean =np.mean(fit_collect,axis=2)
plt.pcolormesh(fc_std)
plt.show()
plt.pcolormesh(fc_mean)
plt.show()